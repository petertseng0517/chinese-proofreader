import hashlib
import hmac
import io
import json
import os

import anthropic
import fitz  # PyMuPDF
import streamlit as st
from anthropic import Anthropic
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

load_dotenv()

APP_USERNAME = os.environ.get("APP_USERNAME")
APP_PASSWORD_HASH = os.environ.get("APP_PASSWORD_HASH")

MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB
MAX_PAGES = 500
MODEL = "claude-opus-4-8"

ERROR_SCHEMA = {
    "type": "object",
    "properties": {
        "errors": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "page": {"type": "integer", "description": "PDF 實際頁面順序，從 1 開始"},
                    "line": {
                        "type": ["integer", "null"],
                        "description": "橫書頁面的行號（依輸入中標示的行號），從 1 開始；若該頁標示為「直書」（無行號），請填 null",
                    },
                    "original": {"type": "string", "description": "原文中的錯別字或錯誤片語"},
                    "suggestion": {"type": "string", "description": "建議修正後的文字"},
                    "reason": {"type": "string", "description": "簡短說明錯誤原因"},
                    "context": {
                        "type": "string",
                        "description": "original 在原文中的上下文片段（original 前後各約 5～10 個字），用於人工定位，尤其當 line 為 null 時必須清楚",
                    },
                },
                "required": ["page", "line", "original", "suggestion", "reason", "context"],
                "additionalProperties": False,
            },
        }
    },
    "required": ["errors"],
    "additionalProperties": False,
}

SYSTEM_PROMPT = (
    "你是專業的繁體中文校稿員。使用者會提供 PDF 逐頁擷取出的文字，請檢查錯別字、錯字、用詞錯誤、明顯的語法錯誤。"
    "只回報你有信心的錯誤，不要回報風格或語氣偏好。"
    "每頁文字前會標示該頁是「橫書」或「直書」：\n"
    "- 橫書頁面：每行前面以「行號:」標示，請將 line 填為對應的行號，不要自行推算或使用印刷頁碼。\n"
    "- 直書頁面：文字已依正確閱讀順序（由右到左、由上到下）排列成連續段落，沒有行號，請將 line 填為 null，"
    "並務必把 context 欄位填寫清楚，讓人工可以用文字搜尋定位到錯誤位置。\n"
    "page 必須完全依照輸入中標示的頁碼回報。"
    "original 欄位必須是文字中可被精確逐字比對到的片段（盡量簡短，幾個字即可）。"
    "context 欄位無論哪種頁面都要填寫，內容是 original 前後各約 5～10 個字的上下文片段。"
    "這是慈濟醫院40周年專刊的校稿工作。請注意：慈濟人文語彙、專有名詞請勿視為錯別字。證嚴上人、慈濟、靜思、功德會等為正確用詞。人名、地名請勿誤判。佛教術語、梵文音譯為正確用詞。若不確定是否為錯字，請勿列入報告。"
)


@st.cache_resource
def get_client() -> Anthropic:
    return Anthropic()


def detect_vertical_page(page: fitz.Page) -> bool:
    """依文字方向向量判斷該頁是否為直書。"""
    vertical_count = 0
    horizontal_count = 0
    for block in page.get_text("dict").get("blocks", []):
        for line in block.get("lines", []):
            dir_x, dir_y = line.get("dir", (1, 0))
            if abs(dir_y) > abs(dir_x):
                vertical_count += 1
            else:
                horizontal_count += 1
    return vertical_count > horizontal_count


def extract_pages(pdf_bytes: bytes) -> list[dict]:
    """回傳每頁的行清單與書寫方向；頁數採 PDF 實際頁面順序。圖片內容不擷取。"""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    try:
        pages = []
        for page in doc:
            text = page.get_text("text")
            lines = text.split("\n")
            pages.append({"lines": lines, "is_vertical": detect_vertical_page(page)})
        return pages
    finally:
        doc.close()


def build_user_message(pages: list[dict]) -> str:
    parts = []
    for page_num, page in enumerate(pages, start=1):
        lines = page["lines"]
        if page["is_vertical"]:
            full_text = "".join(lines)
            parts.append(
                f"=== 第 {page_num} 頁（PDF 實際頁面順序；直書，文字已依正確閱讀順序排列，無行號） ===\n{full_text}"
            )
        else:
            numbered_lines = "\n".join(f"{i}: {line}" for i, line in enumerate(lines, start=1))
            parts.append(f"=== 第 {page_num} 頁（PDF 實際頁面順序；橫書，含行號） ===\n{numbered_lines}")
    return "\n\n".join(parts)


def proofread(pages: list[dict]) -> list[dict] | None:
    """成功時回傳錯誤清單（可能是空清單）；發生例外或被拒絕時回傳 None。"""
    client = get_client()
    try:
        with client.messages.stream(
            model=MODEL,
            max_tokens=64000,
            thinking={"type": "adaptive"},
            output_config={
                "effort": "high",
                "format": {"type": "json_schema", "schema": ERROR_SCHEMA},
            },
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": build_user_message(pages)}],
        ) as stream:
            response = stream.get_final_message()
    except anthropic.AuthenticationError:
        st.error("API 金鑰無效，請確認 .env 中的 ANTHROPIC_API_KEY 是否正確。")
        return None
    except anthropic.PermissionDeniedError:
        st.error("API 金鑰權限不足，請確認金鑰設定。")
        return None
    except anthropic.RateLimitError:
        st.error("已達 API 速率限制，請稍後再試。")
        return None
    except anthropic.BadRequestError as e:
        if "credit balance" in str(e).lower():
            st.error("帳戶額度不足，請至 Anthropic Console 的 Plans & Billing 加值後再試。")
        else:
            st.error(f"請求發生錯誤：{e.message}")
        return None
    except anthropic.APIStatusError as e:
        st.error(f"API 發生錯誤（{e.status_code}）：{e.message}")
        return None

    if response.stop_reason == "refusal":
        st.error("模型拒絕處理此內容，請確認文件內容後重試。")
        return None

    if response.stop_reason == "max_tokens":
        st.error("文件內容過多，校稿結果在 token 上限前被截斷，請將 PDF 拆成較小檔案後重試。")
        return None

    text_block = next((b for b in response.content if b.type == "text"), None)
    if text_block is None:
        st.error("未收到模型回應內容，請重試。")
        return None
    try:
        data = json.loads(text_block.text)
    except json.JSONDecodeError:
        st.error("無法解析模型回應內容，請重試。")
        return None
    errors = data.get("errors", [])
    errors.sort(key=lambda e: (e["page"], e["line"] if e["line"] is not None else 0))
    return errors


def build_report_lines(errors: list[dict]) -> list[str]:
    lines = []
    for e in errors:
        if e.get("line") is not None:
            lines.append(f"第{e['page']}頁，第{e['line']}行，「{e['original']}」→ 建議「{e['suggestion']}」")
        else:
            lines.append(
                f"第{e['page']}頁，上下文「{e['context']}」，「{e['original']}」→ 建議「{e['suggestion']}」"
            )
    return lines


def build_excel_report(errors: list[dict], total_chars: int, error_count: int, error_rate: float) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "校稿結果"

    headers = ["頁碼（PDF實際順序）", "行號", "錯字", "建議修正", "原因", "上下文"]
    ws.append(headers)
    for e in errors:
        ws.append([e["page"], e["line"], e["original"], e["suggestion"], e["reason"], e.get("context", "")])

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 14)

    summary = wb.create_sheet("統計")
    summary.append(["項目", "數值"])
    summary.append(["總字數", total_chars])
    summary.append(["錯字數", error_count])
    summary.append(["錯誤率", f"{error_rate:.2f}%"])
    summary.column_dimensions["A"].width = 14
    summary.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def verify_credentials(username: str, password: str) -> bool:
    if not APP_USERNAME or not APP_PASSWORD_HASH:
        return False
    password_hash = hashlib.sha256(password.encode("utf-8")).hexdigest()
    username_ok = hmac.compare_digest(username, APP_USERNAME)
    password_ok = hmac.compare_digest(password_hash, APP_PASSWORD_HASH)
    return username_ok and password_ok


def render_login() -> None:
    st.title("繁體中文 PDF 校稿工具")
    st.subheader("請先登入")

    if not APP_USERNAME or not APP_PASSWORD_HASH:
        st.error("尚未設定登入帳密，請在 .env 設定 APP_USERNAME 與 APP_PASSWORD_HASH。")
        return

    with st.form("login_form"):
        username = st.text_input("帳號")
        password = st.text_input("密碼", type="password")
        submitted = st.form_submit_button("登入")

    if submitted:
        if verify_credentials(username, password):
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("帳號或密碼錯誤。")


def main():
    st.set_page_config(page_title="繁體中文 PDF 校稿工具", layout="wide")

    if not st.session_state.get("authenticated"):
        render_login()
        return

    with st.sidebar:
        if st.button("登出"):
            st.session_state.clear()
            st.rerun()

    st.title("繁體中文 PDF 校稿工具")
    st.caption("頁數為 PDF 實際頁面順序，非頁面上印刷的頁碼。")
    st.caption("花蓮慈濟醫院公共傳播室專用")

    if not os.environ.get("ANTHROPIC_API_KEY"):
        st.warning("尚未設定環境變數 ANTHROPIC_API_KEY，請先設定後再使用。")

    uploaded_file = st.file_uploader("上傳 PDF 檔案", type=["pdf"])
    if uploaded_file is None:
        return

    pdf_bytes = uploaded_file.getvalue()

    if len(pdf_bytes) > MAX_FILE_SIZE:
        st.error(f"檔案大小超過限制（{MAX_FILE_SIZE // (1024 * 1024)}MB）。")
        return

    pages = extract_pages(pdf_bytes)

    if len(pages) > MAX_PAGES:
        st.error(f"頁數超過限制（{MAX_PAGES} 頁），目前為 {len(pages)} 頁。")
        return

    if st.session_state.get("file_name") != uploaded_file.name:
        with st.spinner("正在校稿，請稍候..."):
            errors = proofread(pages)
        if errors is None:
            # 校稿失敗，清除舊狀態避免顯示上一份檔案的結果
            st.session_state.pop("errors", None)
            st.session_state.pop("file_name", None)
            st.session_state.pop("pages", None)
            return
        st.session_state["errors"] = errors
        st.session_state["file_name"] = uploaded_file.name
        st.session_state["pages"] = pages

    if "errors" not in st.session_state:
        return

    errors = st.session_state["errors"]
    pages = st.session_state["pages"]

    total_chars = sum(len(line) for page in pages for line in page["lines"])
    error_count = len(errors)
    error_rate = (error_count / total_chars * 100) if total_chars else 0.0

    st.subheader("校稿報告")
    report_lines = build_report_lines(errors)
    if report_lines:
        st.text("\n".join(report_lines))
    else:
        st.success("沒有發現錯誤。")

    st.subheader("統計")
    col1, col2, col3 = st.columns(3)
    col1.metric("總字數", total_chars)
    col2.metric("錯字數", error_count)
    col3.metric("錯誤率", f"{error_rate:.2f}%")

    excel_bytes = build_excel_report(errors, total_chars, error_count, error_rate)
    st.download_button(
        "下載校稿報告（Excel）",
        data=excel_bytes,
        file_name=f"proofread_report_{os.path.splitext(uploaded_file.name)[0]}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
